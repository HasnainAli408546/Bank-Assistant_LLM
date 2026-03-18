from openpyxl import load_workbook

file = "NUST Bank-Product-Knowledge.xlsx"
wb = load_workbook(file, data_only=True)

# Adjust this set with all the blue header colors you actually use
BLUE_RGB_VALUES = {
    "FF0070C0"#,  # main blue
    # "FFD8D8D8",  # light gray
    # "FFFEF2CB",  # pale header
}

blue_cells_count = 0
blue_texts = []

for sheet in wb.worksheets:
    for row in sheet.iter_rows():
        for cell in row:
            fill = cell.fill
            if not fill or fill.fill_type != "solid":
                continue

            color_obj = fill.fgColor
            color = getattr(color_obj, "rgb", None)

            if not color or color == "00000000":
                continue

            if color in BLUE_RGB_VALUES:
                blue_cells_count += 1

                if cell.value is not None:
                    blue_texts.append(str(cell.value).strip())

print("Total colored (blue) cells detected:", blue_cells_count)
print()

print("All detected blue cells:\n")

for i, text in enumerate(blue_texts, 1):
    print(f"{i}. {text}")